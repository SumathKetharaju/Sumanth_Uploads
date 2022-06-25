import openpyxl

CURRENT_PATH = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\New_HCL_DSR.xlsx"

current_workbook_obj = openpyxl.load_workbook(CURRENT_PATH)

current_sheet_obj = current_workbook_obj.active

max_rows = current_sheet_obj.max_row
max_columns = current_sheet_obj.max_column

for current_row in range(1, max_rows+1):
    print(f"{'#'*50} current row is --> {current_row} {'#'*50}", "\n")
    for current_column in range(1, max_columns+1):
        cell_obj = current_sheet_obj.cell(row=current_row, column=current_column)
        current_column_data = cell_obj.value
        if current_column_data != None and current_column_data == "date":
            print(cell_obj.value)
            cell_obj.value = current_column_data.replace("date", "Sumanth")
            print(cell_obj.value)
            print(f"data at column {current_column} is --> {current_column_data}", "\n")
            print("-----------------changed-----------")
current_workbook_obj.save(CURRENT_PATH)

