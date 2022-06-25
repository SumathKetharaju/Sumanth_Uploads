# requirement --> from folder we can read the all xlsx files and then those files should be added as a different sheet
# within a single xlsx file
# read data from each xlsx file
import os
import openpyxl

# print(os.listdir())
# print(os.listdir().endswith(".xlsx"))

xlsx_files = [x for x in os.listdir() if x.endswith(".xlsx")]

print(f"List of all xlsx files are --> {xlsx_files}")
print(f"count of xlsx files is --> {len(xlsx_files)}")

PATH = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\Team_DSR.xlsx"

all_columns_data = []
all_rows_data = []

for index in range(len(xlsx_files)):

    print(f"index of the file is --> {index} and name of that file is --> {xlsx_files[index]}")
    current_work_book = openpyxl.load_workbook(xlsx_files[index])
    current_sheet_obj = current_work_book.active
    max_rows_of_current_sheet = current_sheet_obj.max_row
    max_columns_of_current_sheet = current_sheet_obj.max_column
    print(f"max columns of current sheet is --> {max_columns_of_current_sheet}")
    print(f"max rows of current sheet is --> {max_rows_of_current_sheet}", "\n")

    for current_row in range(1, max_rows_of_current_sheet+1):
        print(f"{'#'*50} Now {xlsx_files[index]} is at row number --> {current_row} {'#'*50}", "\n")
        for current_column in range(1, max_columns_of_current_sheet+1):
            # print(f"{'!'*50} Now {xlsx_files[index]} is at column number --> {current_column} {'!'*50}", "\n")
            current_cell_obj = current_sheet_obj.cell(row=current_row, column=current_column)
            each_column_data = current_cell_obj.value
            print(f"{xlsx_files[index]} data at column {current_column} is --> {each_column_data}", "\n")
            all_columns_data.append(each_column_data)
            # print(f"{'.'*50} {xlsx_files[index]} is completed column number {current_column} {'.'*50}", "\n")
        each_row_data = all_columns_data[:]
        # print(f"{xlsx_files[index]} data at row number {current_row} is --> {each_row_data}", "\n")
        all_rows_data.append(each_row_data)
        print(f"{xlsx_files[index]} data at row {current_row} is --> {each_row_data}", "\n")
        all_columns_data.clear()
        # print(f"{'^'*50} {xlsx_files[index]} is completed row number {current_row} {'^'*50}", "\n")
    print(f"{'*'*50} Now {xlsx_files[index]} all rows are completed {'*'*50}", "\n")
    total_data = all_rows_data[:]
    print(f"All the data is '\n' '\n' {total_data}")
    sheet_name = xlsx_files[index].split(".xlsx")[0]
    work_book = openpyxl.Workbook()
    # cuttent_sheet = work_book.active
    sheet = work_book.create_sheet(sheet_name)
    for item in total_data:
        sheet.append(item)
    # work_book.save(PATH)
    all_rows_data.clear()
    break

    # print(f"sheet name is --> {sheet_name}", "\n")

print("--------------------------out of loop------------------------------------")



