import openpyxl
import re

path = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\Students_Details.xlsx"

students_marks = []
students = []
expected_index = 0

workbook_obj = openpyxl.load_workbook(path)

sheet_obj = workbook_obj.active

max_rows = sheet_obj.max_row
max_columns = sheet_obj.max_column

print(f"Maximum rows are {max_rows}", "\n")
print(f"Maximum columns are {max_columns}", "\n")

print("----------------------------------------particular columns data----------------------------------------------")
for row in range(1, max_rows+1):
    cell_obj_1 = sheet_obj.cell(row=row, column=1)
    student = cell_obj_1.value
    if row != 1:
        students.append(student)
    cell_obj = sheet_obj.cell(row=row, column=2)
    Marks = cell_obj.value
    # print(Marks)
    if row != 1:
        students_marks.append(Marks)
# print()

print(students_marks)

max_marks = students_marks[0]
# max_marks = [max_marks for num in students_marks if num > max_marks]
for num in students_marks:
    if num > max_marks:
        max_marks = num

print(f"Maximum Marks are --> {max_marks}")

for mark in students_marks:
    if mark == max_marks:
        expected_index = students_marks.index(mark)
        print(expected_index)

print(students)
# print(students[expected_index])
print(f"Top Scorer Student is --> {students[expected_index]}")


















# print("-----------------------------------------All Data ---------------------------------------------------------")
# for row in range(1, max_rows+1):
#     print(f"----------------------This is Row {row}-------------------------------------")
#     for column in range(1, max_columns+1):
#         cell_obj = sheet_obj.cell(row=row, column=column)
#         marks = cell_obj.value
#         print(f"----------------------This is column {column}-------------------------------------")
#         maximum_marks_pattern = re.compile(str(max_marks))
#         maximum_marks_match = maximum_marks_pattern.search(str(marks))
#
#         if maximum_marks_match:
#             print(marks)
#             print(maximum_marks_match)
#             # print(marks)
#     print(marks)

# print("---------------------------------------particular row data-------------------------------------------------")
#
# for column in range(1, max_columns+1):
#     cell_obj = sheet_obj.cell(row=1, column=column)
#     print(cell_obj.value, end=" ")
# print()




