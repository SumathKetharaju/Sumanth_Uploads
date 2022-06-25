import openpyxl

path = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\HCL_DSR.xlsx"

workbook_obj = openpyxl.load_workbook(path)

sheet_obj = workbook_obj.active

max_rows = sheet_obj.max_row
max_columns = sheet_obj.max_column

print(f"Maximum rows are {max_rows}", "\n")
print(f"Maximum columns are {max_columns}", "\n")

print("-----------------------------------------All Data ------------------------------------------------------------")
for row in range(1, max_rows+1):
    for column in range(1, max_columns+1):
        cell_obj = sheet_obj.cell(row=row, column=column)
        print(cell_obj.value, end=" ")
    print()

print("---------------------------------------particular row data----------------------------------------------------")

for column in range(1, max_columns+1):
    cell_obj = sheet_obj.cell(row=1, column=column)
    print(cell_obj.value, end=" ")
print()

print("----------------------------------------particular columns data----------------------------------------------")
for row in range(1, max_rows+1):
    cell_obj = sheet_obj.cell(row=row, column=2)
    print(cell_obj.value)
print()
