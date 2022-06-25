import openpyxl
import re

CURRENT_PATH = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\Employee_Skill_Profile.xlsx"

current_workbook_obj = openpyxl.load_workbook(CURRENT_PATH)

current_sheet_obj = current_workbook_obj.active

max_rows = current_sheet_obj.max_row
max_columns = current_sheet_obj.max_column

print(f"Maximum rows are --> {max_rows}", "\n")
print(f"Maximum columns are --> {max_columns}", "\n")

all_column_data = []
all_row_data = []

skill_python = "Python"
skill_embeddedc = "Embeddedc"
skill_c = "c++"

python_resources = []
embedded_resources = []
c_resources = []

for current_row in range(1, max_rows+1):
    print(f"{'#'*50} current row is --> {current_row} {'#'*50}", "\n")
    for current_column in range(1, max_columns+1):
        cell_obj = current_sheet_obj.cell(row=current_row, column=current_column)
        current_column_data = cell_obj.value
        all_column_data.append(current_column_data)
        print(f"data at column {current_column} is --> {current_column_data}", "\n")
    current_row_data = all_column_data[:]
    print(f"data at row {current_row} is --> {current_row_data}", "\n")
    all_column_data.clear()
    all_row_data.append(current_row_data)

    if skill_python in current_row_data:
        python_resources.append(current_row_data[0])
    elif skill_embeddedc in current_row_data:
        embedded_resources.append(current_row_data[0])
    elif skill_c in current_row_data:
        c_resources.append(current_row_data[0])

print(f"python_resources are {python_resources}")
print(f"embedded_resources are {embedded_resources}")
print(f"c_resources are {c_resources}")
print("-----------------------------------------------------------------------")
print(f"Number of python_resources are {len(python_resources)}")
print(f"Number of embedded_resources are {len(embedded_resources)}")
print(f"Number of c_resources are {len(c_resources)}")


