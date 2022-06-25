import openpyxl

CURRENT_PATH = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\HCL_DSR.xlsx"
NEW_PATH = r"C:\Users\SUMANTH\Desktop\HCL_Sumanth\New_HCL_DSR.xlsx"

current_workbook_obj = openpyxl.load_workbook(CURRENT_PATH)

current_sheet_obj = current_workbook_obj.active

max_rows = current_sheet_obj.max_row
max_columns = current_sheet_obj.max_column

print(f"Maximum rows are --> {max_rows}", "\n")
print(f"Maximum columns are --> {max_columns}", "\n")
columns_data = []
rows_data = []

for current_row in range(1, max_rows+1):
    print(f"{'#'*50} current row is --> {current_row} {'#'*50}", "\n")
    for current_column in range(1, max_columns+1):
        cell_obj = current_sheet_obj.cell(row=current_row, column=current_column)
        current_column_data = cell_obj.value
        columns_data.append(current_column_data)
        print(f"data at column {current_column} is --> {current_column_data}", "\n")
    current_row_data = columns_data[:]
    print(f"data at row {current_row} is --> {current_row_data}", "\n")
    columns_data.clear()
    rows_data.append(current_row_data)
total_data = rows_data[:]
rows_data.clear()
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
for data in total_data:
    new_sheet.append(data)
new_workbook.save(NEW_PATH)


