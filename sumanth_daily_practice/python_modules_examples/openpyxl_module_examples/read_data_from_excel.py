import openpyxl

path = r"C:\Users\SUMANTH\Desktop\Backup_Folder\Sumanth_DSR.xlsx"

workbook_obj = openpyxl.load_workbook(path)

sheet_obj = workbook_obj.active

max_rows = sheet_obj.max_row
print(max_rows)

max_columns = sheet_obj.max_column
print(max_columns)

for row in range(1, max_rows):
    for column in range(1, max_columns):
        cell_obj = sheet_obj.cell(row=row, column=column)

        print(cell_obj.value, end="")
    print()


