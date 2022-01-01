import openpyxl

path = r"C:\Users\SUMANTH\Desktop\Daily_Status_of_Sumanth_22_to_27.xlsx"

workbook_obj = openpyxl.load_workbook(path)

sheet_obj = workbook_obj.active

max_rows = sheet_obj.max_row
max_columns = sheet_obj.max_column

print("---------------Read All Data from excel file------------------------")

for row in range(1, max_rows+1):
    for column in range(2, max_columns+1):
        cell_obj = sheet_obj.cell(row=row, column=column)
        print(cell_obj.value)
    print()

print("-----------------Read Particular Data from excel file----------------------")

cell_obj_1 = sheet_obj.cell(row=4, column=3)
print(cell_obj_1.value)
print("-----------------Read Particular column Data from excel file----------------------")

for row in range(1, max_rows+1):
    cell_obj = sheet_obj.cell(column=2, row=row)
    print(cell_obj.value)

print("----------------Read Particular Row Data from excel file-----------------------")

for column in range(1, max_rows+1):
    cell_obj = sheet_obj.cell(row=1, column=column)
    print(cell_obj.value)
